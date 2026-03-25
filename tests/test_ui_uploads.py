import io
import unittest

import pandas as pd
from openpyxl import Workbook, load_workbook

from listcompare.core.products.product_mapping import build_product_map
from listcompare.core.products.product_schema import HICORE_COLUMNS
from listcompare.core.suppliers.profile import build_supplier_hicore_renamed_copy
from listcompare.interfaces.ui.io.exports import _df_excel_bytes
from listcompare.interfaces.ui.io.index_names import _load_names_from_uploaded_hicore
from listcompare.interfaces.ui.io.uploads import (
    _normalize_hicore_identifier_columns,
    _read_hicore_upload,
    _read_supplier_upload,
)
from listcompare.interfaces.ui.services.compare_pipeline import load_hicore_compare_df


def _xlsx_bytes_from_rows(rows: list[list[object]], *, number_formats: dict[str, str]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    for row in rows:
        worksheet.append(row)

    for cell_ref, number_format in number_formats.items():
        worksheet[cell_ref].number_format = number_format

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _hicore_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(sep=";", index=False).encode("utf-8-sig")


class UiUploadTests(unittest.TestCase):
    def test_read_supplier_upload_preserves_zero_padded_numeric_excel_cells(self) -> None:
        upload_bytes = _xlsx_bytes_from_rows(
            [
                ["SupplierSku", "SupplierArticleNo", "Price"],
                [123, 45, 99.95],
            ],
            number_formats={
                "A2": "000000",
                "B2": "0000",
                "C2": "0.00",
            },
        )

        df_supplier = _read_supplier_upload("supplier.xlsx", upload_bytes)

        self.assertEqual(df_supplier["SupplierSku"].tolist(), ["000123"])
        self.assertEqual(df_supplier["SupplierArticleNo"].tolist(), ["0045"])
        self.assertEqual(df_supplier["Price"].tolist(), ["99.95"])

    def test_supplier_transform_checkbox_respects_zero_padded_excel_upload_values(self) -> None:
        upload_bytes = _xlsx_bytes_from_rows(
            [
                ["SupplierSku", "NameCol"],
                [123, "Product A"],
            ],
            number_formats={"A2": "000000"},
        )
        df_supplier = _read_supplier_upload("supplier.xlsx", upload_bytes)

        renamed_keep = build_supplier_hicore_renamed_copy(
            df_supplier,
            target_to_source={
                "Art.märkning": "SupplierSku",
                "Artikelnamn": "NameCol",
            },
            supplier_name="EM Nordic",
            strip_leading_zeros_from_sku=False,
        )
        renamed_strip = build_supplier_hicore_renamed_copy(
            df_supplier,
            target_to_source={
                "Art.märkning": "SupplierSku",
                "Artikelnamn": "NameCol",
            },
            supplier_name="EM Nordic",
            strip_leading_zeros_from_sku=True,
        )

        self.assertEqual(renamed_keep["Art.märkning"].tolist(), ["000123"])
        self.assertEqual(renamed_strip["Art.märkning"].tolist(), ["123"])

        workbook = load_workbook(
            io.BytesIO(_df_excel_bytes(renamed_keep, sheet_name="HiCore-format")),
            data_only=False,
        )
        sheet = workbook.active
        self.assertEqual(sheet["A1"].value, "Art.märkning")
        self.assertEqual(sheet["A2"].value, "000123")
        self.assertEqual(sheet["A2"].number_format, "@")

    def test_hicore_csv_load_and_product_map_preserve_leading_zeroes(self) -> None:
        sku_col = HICORE_COLUMNS["sku"]
        name_col = HICORE_COLUMNS["name"]
        stock_col = HICORE_COLUMNS["stock"]
        price_col = HICORE_COLUMNS["price"]
        supplier_col = HICORE_COLUMNS["supplier"]
        df_hicore = pd.DataFrame(
            [
                {
                    sku_col: "000123",
                    name_col: "Product A",
                    stock_col: "1",
                    price_col: "10",
                    supplier_col: "EM Nordic",
                }
            ]
        )

        loaded_df = load_hicore_compare_df("hicore.csv", _hicore_csv_bytes(df_hicore))
        hicore_map = build_product_map(
            loaded_df,
            source="hicore",
            columns=HICORE_COLUMNS,
        )

        self.assertEqual(loaded_df[sku_col].tolist(), ["000123"])
        self.assertEqual(list(hicore_map.keys()), ["000123"])
        self.assertEqual(hicore_map["000123"][0].sku, "000123")

    def test_hicore_excel_load_and_product_map_preserve_leading_zeroes(self) -> None:
        sku_col = HICORE_COLUMNS["sku"]
        name_col = HICORE_COLUMNS["name"]
        stock_col = HICORE_COLUMNS["stock"]
        price_col = HICORE_COLUMNS["price"]
        supplier_col = HICORE_COLUMNS["supplier"]
        upload_bytes = _xlsx_bytes_from_rows(
            [
                [sku_col, name_col, stock_col, price_col, supplier_col],
                [123, "Product A", 1, 10, "EM Nordic"],
            ],
            number_formats={"A2": "000000"},
        )

        loaded_df = load_hicore_compare_df("hicore.xlsx", upload_bytes)
        hicore_map = build_product_map(
            loaded_df,
            source="hicore",
            columns=HICORE_COLUMNS,
        )

        self.assertEqual(loaded_df[sku_col].tolist(), ["000123"])
        self.assertEqual(list(hicore_map.keys()), ["000123"])
        self.assertEqual(hicore_map["000123"][0].sku, "000123")

    def test_read_hicore_upload_preserves_zero_padded_excel_cells(self) -> None:
        sku_col = HICORE_COLUMNS["sku"]
        upload_bytes = _xlsx_bytes_from_rows(
            [
                [sku_col, "Webbordernr"],
                [123, 24541],
            ],
            number_formats={
                "A2": "000000",
                "B2": "000000000",
            },
        )

        df_hicore = _read_hicore_upload("hicore.xlsx", upload_bytes)

        self.assertEqual(df_hicore[sku_col].tolist(), ["000123"])
        self.assertEqual(df_hicore["Webbordernr"].tolist(), ["000024541"])

    def test_read_hicore_upload_uses_excel_sheet_with_hicore_headers(self) -> None:
        sku_col = HICORE_COLUMNS["sku"]
        supplier_col = HICORE_COLUMNS["supplier"]
        workbook = Workbook()
        wrong_sheet = workbook.active
        wrong_sheet.title = "Info"
        wrong_sheet.append(["NotHiCoreColumn"])
        wrong_sheet.append(["ignore"])

        data_sheet = workbook.create_sheet("Products")
        data_sheet.append([sku_col, supplier_col])
        data_sheet.append([123, "EM Nordic"])
        data_sheet["A2"].number_format = "000000"

        buffer = io.BytesIO()
        workbook.save(buffer)
        upload_bytes = buffer.getvalue()

        df_hicore = _read_hicore_upload("hicore.xlsx", upload_bytes)

        self.assertEqual(df_hicore[sku_col].tolist(), ["000123"])
        self.assertEqual(df_hicore[supplier_col].tolist(), ["EM Nordic"])

    def test_load_names_from_uploaded_hicore_reads_excel_upload(self) -> None:
        sku_col = HICORE_COLUMNS["sku"]
        supplier_col = HICORE_COLUMNS["supplier"]
        brand_col = HICORE_COLUMNS["brand"]
        workbook = Workbook()
        wrong_sheet = workbook.active
        wrong_sheet.title = "Info"
        wrong_sheet.append(["Ignore"])

        data_sheet = workbook.create_sheet("Products")
        data_sheet.append([sku_col, supplier_col, brand_col])
        data_sheet.append([123, "EM Nordic", "Sony"])
        data_sheet["A2"].number_format = "000000"

        buffer = io.BytesIO()
        workbook.save(buffer)
        upload_bytes = buffer.getvalue()

        supplier_names, brand_names, has_supplier_column, has_brand_column = (
            _load_names_from_uploaded_hicore("hicore.xlsx", upload_bytes)
        )

        self.assertEqual(supplier_names, ["EM Nordic"])
        self.assertEqual(brand_names, ["Sony"])
        self.assertTrue(has_supplier_column)
        self.assertTrue(has_brand_column)

    def test_normalize_hicore_identifier_columns_strips_integer_like_decimal_suffixes(self) -> None:
        sku_col = HICORE_COLUMNS["sku"]
        article_number_col = HICORE_COLUMNS["article_number"]
        df_hicore = pd.DataFrame(
            [
                {
                    sku_col: "123.0",
                    article_number_col: "456.0",
                    "Webbordernr": "789.0",
                    HICORE_COLUMNS["name"]: "Product A",
                }
            ]
        )

        normalized_df = _normalize_hicore_identifier_columns(df_hicore)

        self.assertEqual(normalized_df[sku_col].tolist(), ["123"])
        self.assertEqual(normalized_df[article_number_col].tolist(), ["456"])
        self.assertEqual(normalized_df["Webbordernr"].tolist(), ["789"])
        self.assertEqual(normalized_df[HICORE_COLUMNS["name"]].tolist(), ["Product A"])


if __name__ == "__main__":
    unittest.main()
