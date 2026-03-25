from __future__ import annotations

import csv
import io
import re
from pathlib import Path
from typing import Optional

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from ....core.products.product_schema import HICORE_COLUMNS
from ..common import CSV_ENCODINGS


def _uploaded_csv_to_df(
    data: bytes,
    *,
    sep: str | None,
    engine: Optional[str] = None,
    extra_read_csv_kwargs: Optional[dict[str, object]] = None,
) -> pd.DataFrame:
    last_err: Optional[Exception] = None
    for enc in CSV_ENCODINGS:
        try:
            text = data.decode(enc)
            kwargs: dict[str, object] = {
                "sep": sep,
                "dtype": str,
                "index_col": False,
            }
            if engine is not None:
                kwargs["engine"] = engine
            if extra_read_csv_kwargs:
                kwargs.update(extra_read_csv_kwargs)
            return pd.read_csv(io.StringIO(text), **kwargs)
        except UnicodeDecodeError as err:
            last_err = err
        except Exception as err:
            last_err = err
    if last_err is not None:
        raise last_err
    raise UnicodeDecodeError("utf-8", b"", 0, 1, "Unable to decode uploaded CSV")


def _read_supplier_csv_upload(data: bytes) -> pd.DataFrame:
    try:
        return _uploaded_csv_to_df(data, sep=None, engine="python")
    except Exception as first_error:
        fallback_error: Exception = first_error

        for sep in (";", ",", "\t", "|"):
            try:
                return _uploaded_csv_to_df(data, sep=sep, engine="python")
            except Exception as err:
                fallback_error = err

        for sep in (";", ",", "\t", "|"):
            try:
                return _uploaded_csv_to_df(
                    data,
                    sep=sep,
                    engine="python",
                    extra_read_csv_kwargs={
                        # Fallback for malformed CSV quotes in some supplier exports.
                        "quoting": csv.QUOTE_NONE,
                    },
                )
            except Exception as err:
                fallback_error = err

        raise ValueError(
            "Kunde inte l\u00e4sa CSV-filen. Filen verkar inneh\u00e5lla trasig CSV-formatering "
            f"(t.ex. citattecken). Originalfel: {first_error}. Senaste fallback-fel: {fallback_error}"
        ) from first_error


def _read_compare_magento_csv_upload(data: bytes) -> pd.DataFrame:
    return _read_supplier_csv_upload(data)


def _read_hicore_csv_upload(data: bytes) -> pd.DataFrame:
    return _uploaded_csv_to_df(data, sep=";")


def _find_case_insensitive_column(columns: list[object], wanted: str) -> Optional[str]:
    wanted_folded = str(wanted).strip().casefold()
    for column in columns:
        if str(column).strip().casefold() == wanted_folded:
            return str(column)
    return None


def _normalize_integer_like_identifier_text(value: object) -> object:
    if pd.isna(value):
        return value
    text = str(value).strip()
    if re.fullmatch(r"-?\d+\.0+", text) is None:
        return value
    return text.split(".", 1)[0]


def _normalize_hicore_identifier_columns(df: pd.DataFrame) -> pd.DataFrame:
    normalized_df = df.copy()
    for wanted_column in (
        HICORE_COLUMNS["sku"],
        HICORE_COLUMNS["article_number"],
        "Webbordernr",
    ):
        actual_column = _find_case_insensitive_column(normalized_df.columns.tolist(), wanted_column)
        if actual_column is None:
            continue
        normalized_df[actual_column] = normalized_df[actual_column].map(
            _normalize_integer_like_identifier_text
        )
    return normalized_df


def _zero_pad_width_from_excel_number_format(number_format: object) -> Optional[int]:
    text = str(number_format or "").strip()
    if text == "":
        return None

    primary_section = text.split(";", 1)[0]
    primary_section = re.sub(r'"[^"]*"', "", primary_section)
    primary_section = re.sub(r"\[[^\]]*\]", "", primary_section)
    primary_section = re.sub(r"\\.", "", primary_section)
    primary_section = primary_section.strip()
    if re.fullmatch(r"0+", primary_section) is None:
        return None
    return len(primary_section)


def _formatted_zero_padded_excel_text(raw_value: object, number_format: object) -> Optional[str]:
    width = _zero_pad_width_from_excel_number_format(number_format)
    if width is None or isinstance(raw_value, bool):
        return None
    if isinstance(raw_value, int):
        integer_value = raw_value
    elif isinstance(raw_value, float):
        if not raw_value.is_integer():
            return None
        integer_value = int(raw_value)
    else:
        return None

    sign = "-" if integer_value < 0 else ""
    digits = str(abs(integer_value)).zfill(width)
    return f"{sign}{digits}"


def _matching_column_count(columns: list[object], wanted_columns: list[str]) -> int:
    actual = {str(column).strip().casefold() for column in columns}
    return sum(1 for column in wanted_columns if column.casefold() in actual)


def _best_hicore_sheet_name(data: bytes) -> Optional[str]:
    wanted_columns = [
        str(column).strip()
        for column in HICORE_COLUMNS.values()
        if isinstance(column, str) and str(column).strip() != ""
    ]
    workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    try:
        best_sheet_name: Optional[str] = None
        best_score = -1
        sku_column_name = HICORE_COLUMNS["sku"]
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            header_cells = next(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
                (),
            )
            header_values = [str(value).strip() for value in header_cells if value is not None]
            if _find_case_insensitive_column(header_values, sku_column_name) is None:
                continue
            score = _matching_column_count(header_values, wanted_columns)
            if score > best_score:
                best_score = score
                best_sheet_name = sheet_name
        return best_sheet_name
    finally:
        workbook.close()


def _read_excel_upload(data: bytes, *, sheet_name: Optional[str] = None) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(data), dtype=str, sheet_name=sheet_name or 0)
    workbook = load_workbook(io.BytesIO(data), data_only=True)
    try:
        worksheet_name = sheet_name or workbook.sheetnames[0]
        worksheet = workbook[worksheet_name]
        repaired_df = df.copy()
        for row_index in range(len(repaired_df.index)):
            for column_index in range(len(repaired_df.columns)):
                cell = worksheet.cell(row=row_index + 2, column=column_index + 1)
                repaired_value = _formatted_zero_padded_excel_text(
                    cell.value,
                    cell.number_format,
                )
                if repaired_value is None:
                    continue
                repaired_df.iat[row_index, column_index] = repaired_value
        return repaired_df
    finally:
        workbook.close()


def _read_supplier_upload(file_name: str, data: bytes) -> pd.DataFrame:
    return _read_supplier_upload_cached(file_name=file_name, data=data).copy()


def _read_hicore_upload(file_name: str, data: bytes) -> pd.DataFrame:
    return _read_hicore_upload_cached(file_name=file_name, data=data).copy()


def _read_hicore_name_columns(
    file_name: str,
    data: bytes,
) -> tuple[list[str], list[str], bool, bool]:
    supplier_names, brand_names, has_supplier_column, has_brand_column = (
        _read_hicore_name_columns_cached(file_name=file_name, data=data)
    )
    return (
        list(supplier_names),
        list(brand_names),
        has_supplier_column,
        has_brand_column,
    )


@st.cache_data(show_spinner=False)
def _read_supplier_upload_cached(file_name: str, data: bytes) -> pd.DataFrame:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv":
        return _read_supplier_csv_upload(data)
    if suffix in (".xlsx", ".xls", ".xlsm"):
        return _read_excel_upload(data)
    raise ValueError(f"Unsupported supplier file type: {file_name}")


@st.cache_data(show_spinner=False)
def _read_hicore_upload_cached(file_name: str, data: bytes) -> pd.DataFrame:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv":
        return _normalize_hicore_identifier_columns(_read_hicore_csv_upload(data))
    if suffix in (".xlsx", ".xls", ".xlsm"):
        sheet_name = _best_hicore_sheet_name(data)
        if sheet_name is None:
            raise ValueError(
                f'Kunde inte hitta ett Excel-blad med kolumnen "{HICORE_COLUMNS["sku"]}" i HiCore-filen.'
            )
        return _normalize_hicore_identifier_columns(
            _read_excel_upload(data, sheet_name=sheet_name)
        )
    raise ValueError(f"Unsupported HiCore file type: {file_name}")


def _raw_text_or_empty(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def _extract_hicore_name_columns_from_excel(
    data: bytes,
    *,
    sheet_name: str,
) -> tuple[list[str], list[str], bool, bool]:
    workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    try:
        worksheet = workbook[sheet_name]
        header_cells = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
            (),
        )
        header_values = [_raw_text_or_empty(value) for value in header_cells]
        supplier_column_name = HICORE_COLUMNS["supplier"]
        brand_column_name = HICORE_COLUMNS.get("brand")
        supplier_index: Optional[int] = None
        brand_index: Optional[int] = None
        for index, header_value in enumerate(header_values):
            if supplier_index is None and (
                _find_case_insensitive_column([header_value], supplier_column_name) is not None
            ):
                supplier_index = index
            if (
                brand_index is None
                and brand_column_name is not None
                and _find_case_insensitive_column([header_value], brand_column_name) is not None
            ):
                brand_index = index

        supplier_names: list[str] = []
        brand_names: list[str] = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_values = list(row)
            if supplier_index is not None and supplier_index < len(row_values):
                supplier_value = _raw_text_or_empty(row_values[supplier_index])
                if supplier_value != "":
                    supplier_names.append(supplier_value)
            if brand_index is not None and brand_index < len(row_values):
                brand_value = _raw_text_or_empty(row_values[brand_index])
                if brand_value != "":
                    brand_names.append(brand_value)

        return (
            supplier_names,
            brand_names,
            supplier_index is not None,
            brand_index is not None,
        )
    finally:
        workbook.close()


@st.cache_data(show_spinner=False)
def _read_hicore_name_columns_cached(
    file_name: str,
    data: bytes,
) -> tuple[list[str], list[str], bool, bool]:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv":
        df_hicore = _read_hicore_csv_upload(data)
        supplier_col = HICORE_COLUMNS["supplier"]
        brand_col = HICORE_COLUMNS.get("brand")
        supplier_names = (
            [_raw_text_or_empty(value) for value in df_hicore[supplier_col].tolist()]
            if supplier_col in df_hicore.columns
            else []
        )
        brand_names = (
            [_raw_text_or_empty(value) for value in df_hicore[brand_col].tolist()]
            if brand_col and brand_col in df_hicore.columns
            else []
        )
        return (
            [value for value in supplier_names if value != ""],
            [value for value in brand_names if value != ""],
            supplier_col in df_hicore.columns,
            bool(brand_col and brand_col in df_hicore.columns),
        )
    if suffix in (".xlsx", ".xls", ".xlsm"):
        sheet_name = _best_hicore_sheet_name(data)
        if sheet_name is None:
            raise ValueError(
                f'Kunde inte hitta ett Excel-blad med kolumnen "{HICORE_COLUMNS["sku"]}" i HiCore-filen.'
            )
        return _extract_hicore_name_columns_from_excel(data, sheet_name=sheet_name)
    raise ValueError(f"Unsupported HiCore file type: {file_name}")
